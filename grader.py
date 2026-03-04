"""
Football Excel Exam Auto-Grader
Compares student submissions against the solution file and fills in grading cells.

Grading cells receive a 0.0–1.0 percentage so the existing formula =score%*max_pts
computes partial credit automatically.
"""

import shutil
import subprocess
import tempfile
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

# ── paths ──────────────────────────────────────────────────────────────────
SOLUTION_PATH = Path(__file__).parent / "solution.xlsx"
SCRIPTS_DIR   = Path(__file__).parent / "scripts"

# ── colour helpers ──────────────────────────────────────────────────────────
GREEN  = PatternFill("solid", start_color="C6EFCE")
YELLOW = PatternFill("solid", start_color="FFEB9C")
RED    = PatternFill("solid", start_color="FFC7CE")
GREEN_FONT  = Font(color="276221", bold=True)
YELLOW_FONT = Font(color="9C5700", bold=True)
RED_FONT    = Font(color="9C0006", bold=True)

NUM_TOL = 0.02   # 2% relative tolerance


def _eq_val(a, b):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        fa, fb = float(a), float(b)
        if fb == 0:
            return abs(fa) < 1e-6
        return abs(fa - fb) / abs(fb) <= NUM_TOL
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def _col_values(ws, col, row_start, row_end):
    return [ws.cell(r, col).value for r in range(row_start, row_end + 1)]


def _rect_values(ws, row_start, row_end, col_start, col_end):
    return [
        [ws.cell(r, c).value for c in range(col_start, col_end + 1)]
        for r in range(row_start, row_end + 1)
    ]


def _match_rate(student_vals, sol_vals):
    """Positional match rate 0.0–1.0."""
    if not student_vals or not sol_vals:
        return 0.0
    if isinstance(student_vals[0], list):
        flat_s = [v for row in student_vals for v in row]
        flat_r = [v for row in sol_vals     for v in row]
    else:
        flat_s, flat_r = list(student_vals), list(sol_vals)
    pairs = list(zip(flat_s, flat_r))
    if not pairs:
        return 0.0
    return sum(_eq_val(a, b) for a, b in pairs) / len(pairs)


def _set_match_rate(student_vals, sol_vals):
    """
    Order-independent match for dynamic arrays (FILTER/VSTACK).
    Returns fraction of solution values found anywhere in student values.
    """
    def norm(v):
        if v is None:
            return None
        try:
            return round(float(v), 2)
        except (TypeError, ValueError):
            return str(v).strip().lower()

    flat_s = [norm(v) for row in student_vals for v in row if v is not None]
    flat_r = [norm(v) for row in sol_vals     for v in row if v is not None]
    if not flat_r:
        return 0.0
    stu_pool = list(flat_s)
    matched  = 0
    for val in flat_r:
        if val in stu_pool:
            matched += 1
            stu_pool.remove(val)
    return matched / len(flat_r)


def _write_score(ws, row, col, rate, is_manual=False):
    """Write 0.0–1.0 into a grading cell with colour coding."""
    cell = ws.cell(row, col)
    if is_manual:
        cell.value = None
        cell.fill  = YELLOW
        cell.font  = YELLOW_FONT
    else:
        cell.value = round(rate, 4)
        if rate >= 0.8:
            cell.fill = GREEN;  cell.font = GREEN_FONT
        elif rate >= 0.4:
            cell.fill = YELLOW; cell.font = YELLOW_FONT
        else:
            cell.fill = RED;    cell.font = RED_FONT


def recalculate(path: Path) -> Path:
    tmp = Path(tempfile.mkdtemp()) / path.name
    shutil.copy(path, tmp)
    try:
        result = subprocess.run(
            ["python3", str(SCRIPTS_DIR / "recalc.py"), str(tmp)],
            capture_output=True, text=True, timeout=60,
            cwd=str(SCRIPTS_DIR.parent)
        )
        if result.returncode != 0:
            print(f"[WARN] recalc: {result.stderr[:200]}")
    except Exception as e:
        print(f"[WARN] recalc failed: {e}")
    return tmp


def _check_named_ranges(wb, expected_names):
    defined = {n.lower() for n in wb.defined_names.keys()}
    missing = [n for n in expected_names if n.lower() not in defined]
    found   = len(expected_names) - len(missing)
    rate    = found / len(expected_names)
    if not missing:
        return 1.0, "✔ All named ranges found"
    return rate, f"✘ {found}/{len(expected_names)} – Missing: {', '.join(missing)}"


def _check_table_exists(wb, table_name):
    for ws in wb.worksheets:
        for tbl in ws.tables.values():
            if tbl.name.lower() == table_name.lower():
                return 1.0, f"✔ Table '{table_name}' found"
    return 0.0, f"✘ Table '{table_name}' not found"


def grade_file(student_path: Path, recalc: bool = True) -> dict:
    if recalc:
        calc_path = recalculate(student_path)
        sol_path  = recalculate(SOLUTION_PATH)
    else:
        calc_path = student_path
        sol_path  = SOLUTION_PATH

    sol_wb_d = load_workbook(str(sol_path),  data_only=True)
    stu_wb_d = load_workbook(str(calc_path), data_only=True)
    stu_wb   = load_workbook(str(calc_path))  # write scores onto recalculated copy

    results = {}

    # ════════════════════════════════════════════════════════════════════
    # SECTION 1  –  grading col Y (25), rows 3–13
    # ════════════════════════════════════════════════════════════════════
    s1_sol = sol_wb_d["Section 1 "]
    # fix .get for openpyxl workbooks
    def _ws(wb, name):
        return wb[name] if name in wb.sheetnames else None

    s1_stu = _ws(stu_wb_d, "Section 1 ")
    s1_out = _ws(stu_wb,   "Section 1 ")

    S1 = {}
    rate0, det0 = _check_named_ranges(stu_wb_d,
        ["PlayerName","Goals","Assists","Matches","Salary",
         "MarketValue","YellowCards","RedCards","Position"])
    S1["Q0"] = {"max":1, "rate":rate0, "detail":det0}

    if s1_stu:
        for label, max_pts, col in [
            ("Q1",1,12),("Q2",2,13),("Q3",2,14),("Q4",2,15),("Q5",5,16),
            ("Q6",2,17),("Q7",5,18),("Q8",5,19),("Q9",5,20),("Q10",5,21)]:
            r = _match_rate(_col_values(s1_stu,col,16,27), _col_values(s1_sol,col,16,27))
            S1[label] = {"max":max_pts, "rate":r,
                         "detail":f"{'✔' if r>=0.8 else '✘'} {int(r*100)}% cells correct"}
    else:
        for label, m in [("Q1",1),("Q2",2),("Q3",2),("Q4",2),("Q5",5),
                          ("Q6",2),("Q7",5),("Q8",5),("Q9",5),("Q10",5)]:
            S1[label] = {"max":m, "rate":0.0, "detail":"✘ Sheet not found"}

    results["Section 1"] = S1
    if s1_out:
        for i, q in enumerate(["Q0","Q1","Q2","Q3","Q4","Q5","Q6","Q7","Q8","Q9","Q10"]):
            _write_score(s1_out, 3+i, 25, S1[q]["rate"])

    # ════════════════════════════════════════════════════════════════════
    # SECTION 2  –  grading col O (15), rows 3–12
    # ════════════════════════════════════════════════════════════════════
    s2_sol = sol_wb_d["Section 2"]
    s2_stu = _ws(stu_wb_d, "Section 2")
    s2_out = _ws(stu_wb,   "Section 2")

    S2 = {}
    if s2_stu:
        # Q1 – Team/League/Stadium (cols C-E, rows 15-22)
        r = _match_rate(_rect_values(s2_stu,15,22,3,5), _rect_values(s2_sol,15,22,3,5))
        S2["Q1"] = {"max":1, "rate":r, "detail":f"{'✔' if r>=0.8 else '✘'} {int(r*100)}%"}

        # Q2 – Capacity col F
        r = _match_rate(_col_values(s2_stu,6,15,22), _col_values(s2_sol,6,15,22))
        S2["Q2"] = {"max":2, "rate":r, "detail":f"{'✔' if r>=0.8 else '✘'} {int(r*100)}%"}

        # Q3 – Ticket Price col G
        r = _match_rate(_col_values(s2_stu,7,15,22), _col_values(s2_sol,7,15,22))
        S2["Q3"] = {"max":2, "rate":r, "detail":f"{'✔' if r>=0.8 else '✘'} {int(r*100)}%"}

        # Q4 – Coach + Revenue cols H-I
        r = _match_rate(_rect_values(s2_stu,15,22,8,9), _rect_values(s2_sol,15,22,8,9))
        S2["Q4"] = {"max":1, "rate":r, "detail":f"{'✔' if r>=0.8 else '✘'} {int(r*100)}%"}

        # Q5 – Attendance % col J
        r = _match_rate(_col_values(s2_stu,10,15,22), _col_values(s2_sol,10,15,22))
        S2["Q5"] = {"max":1, "rate":r, "detail":f"{'✔' if r>=0.8 else '✘'} {int(r*100)}%"}

        # Q6 – Drop-downs
        try:
            has_dv = len(list(s2_stu.data_validations.dataValidation)) >= 2
            S2["Q6"] = {"max":1, "rate":1.0 if has_dv else 0.0,
                        "detail":"✔ Drop-downs detected" if has_dv else "✘ Drop-downs missing"}
        except Exception:
            S2["Q6"] = {"max":1, "rate":0.0, "detail":"⚠ Could not check"}

        # Q7 – manual
        S2["Q7"] = {"max":5, "rate":None, "detail":"⚠ MANUAL REVIEW – conditional formatting"}

        # Q8 – FILTER+CHOOSECOLS: formula in B32, results spill B32:C39
        # Compare set of (team_name, revenue) pairs
        sv8 = _rect_values(s2_stu, 32, 39, 2, 3)
        rv8 = _rect_values(s2_sol, 32, 39, 2, 3)
        r8  = _set_match_rate(sv8, rv8)
        S2["Q8"] = {"max":5, "rate":r8,
                    "detail":f"{'✔' if r8>=0.8 else '✘'} {int(r8*100)}% (FILTER+CHOOSECOLS, order-independent)"}

        # Q9 – LET+FILTER: formula in E32, results spill E32:F35 (4 rows)
        # Extend range to catch different-length results
        sv9 = _rect_values(s2_stu, 32, 39, 5, 6)
        rv9 = _rect_values(s2_sol, 32, 35, 5, 6)
        r9  = _set_match_rate(sv9, rv9)
        S2["Q9"] = {"max":5, "rate":r9,
                    "detail":f"{'✔' if r9>=0.8 else '✘'} {int(r9*100)}% (LET+FILTER, order-independent)"}

        # Q10 – VSTACK: formula in H32, results spill H32:I35 (4 rows)
        sv10 = _rect_values(s2_stu, 32, 39, 8, 9)
        rv10 = _rect_values(s2_sol, 32, 35, 8, 9)
        r10  = _set_match_rate(sv10, rv10)
        S2["Q10"] = {"max":5, "rate":r10,
                     "detail":f"{'✔' if r10>=0.8 else '✘'} {int(r10*100)}% (VSTACK, order-independent)"}

        # Q11-Q14 – multi-criteria, col E rows 47-50
        for label, max_pts, row in [("Q11",2,47),("Q12",5,48),("Q13",5,49),("Q14",5,50)]:
            sv_v = s2_stu.cell(row,5).value
            rv_v = s2_sol.cell(row,5).value
            r    = 1.0 if _eq_val(sv_v, rv_v) else 0.0
            S2[label] = {"max":max_pts, "rate":r,
                         "detail":f"{'✔' if r else '✘'} Student={sv_v} | Expected={rv_v}"}
    else:
        for q, m in [("Q1",1),("Q2",2),("Q3",2),("Q4",1),("Q5",1),("Q6",1),
                     ("Q7",5),("Q8",5),("Q9",5),("Q10",5),
                     ("Q11",2),("Q12",5),("Q13",5),("Q14",5)]:
            S2[q] = {"max":m, "rate":0.0, "detail":"✘ Sheet not found"}

    results["Section 2"] = S2
    if s2_out:
        for i, q in enumerate(["Q1","Q2","Q3","Q4","Q5","Q6","Q7","Q8","Q9","Q10"]):
            rate = S2[q]["rate"]
            _write_score(s2_out, 3+i, 15, rate if rate is not None else 0.0,
                         is_manual=(rate is None))
        for i, q in enumerate(["Q11","Q12","Q13","Q14"]):
            _write_score(s2_out, 47+i, 10, S2[q]["rate"])

    # ════════════════════════════════════════════════════════════════════
    # SECTION 3  –  grading col C (3), rows 31–40
    # ════════════════════════════════════════════════════════════════════
    s3_sol = sol_wb_d["Section 3"]
    s3_stu = _ws(stu_wb_d, "Section 3")
    s3_out = _ws(stu_wb,   "Section 3")

    S3 = {}
    if s3_stu:
        for label, max_pts, col in [
            ("Q1",1,7),("Q2",1,8),("Q3",2,9),("Q4",2,10),("Q5",2,11),
            ("Q8",5,14),("Q9",5,15),("Q10",5,16)]:
            r = _match_rate(_col_values(s3_stu,col,17,26), _col_values(s3_sol,col,17,26))
            S3[label] = {"max":max_pts, "rate":r,
                         "detail":f"{'✔' if r>=0.8 else '✘'} {int(r*100)}% cells correct"}

        # Q6 – Age: DATEDIF(birthdate, TODAY(), "Y")
        # Compute expected ages directly from birth dates so we never depend on
        # which date the solution file was last saved / recalculated.
        from datetime import date
        today = date.today()
        expected_ages = []
        for r in range(17, 27):
            bday = s3_stu.cell(r, 5).value   # col E = Birth Date
            if bday is not None:
                try:
                    bd = bday.date() if hasattr(bday, "date") else bday
                    age = today.year - bd.year - ((today.month, today.day) < (bd.month, bd.day))
                    expected_ages.append(age)
                except Exception:
                    expected_ages.append(None)
            else:
                expected_ages.append(None)
        student_ages = _col_values(s3_stu, 12, 17, 26)
        # Allow ±1 tolerance (birthday edge cases)
        def _age_ok(s, e):
            if s is None or e is None:
                return False
            try:
                return abs(int(float(s)) - int(e)) <= 1
            except (TypeError, ValueError):
                return False
        if expected_ages:
            matched = sum(_age_ok(s, e) for s, e in zip(student_ages, expected_ages))
            rate6   = matched / len(expected_ages)
        else:
            rate6 = 0.0
        S3["Q6"] = {"max":5, "rate":rate6,
                    "detail":f"{'✔' if rate6>=0.8 else '✘'} {int(rate6*100)}% ages correct (computed from birth dates)"}

        # Q7 – Full Name: PROPER(LastName) & ", " & PROPER(FirstName)
        # Build expected values directly from first/last name columns in the student file
        # so we don't depend on whether the student's formula recalculated.
        expected_names = []
        for r in range(17, 27):
            first = s3_stu.cell(r, 2).value   # col B = First Name
            last  = s3_stu.cell(r, 3).value   # col C = Last Name
            if first is not None and last is not None:
                expected_names.append(f"{str(last).strip().title()}, {str(first).strip().title()}")
            else:
                expected_names.append(None)
        student_names = _col_values(s3_stu, 13, 17, 26)
        def _name_ok(s, e):
            if s is None or e is None:
                return False
            return str(s).strip().lower() == str(e).strip().lower()
        if expected_names:
            matched7 = sum(_name_ok(s, e) for s, e in zip(student_names, expected_names))
            rate7    = matched7 / len(expected_names)
        else:
            rate7 = 0.0
        S3["Q7"] = {"max":2, "rate":rate7,
                    "detail":f"{'✔' if rate7>=0.8 else '✘'} {int(rate7*100)}% full names correct (LastName, FirstName format)"}
    else:
        for q, m in [("Q1",1),("Q2",1),("Q3",2),("Q4",2),("Q5",2),("Q6",5),
                     ("Q7",2),("Q8",5),("Q9",5),("Q10",5)]:
            S3[q] = {"max":m, "rate":0.0, "detail":"✘ Sheet not found"}

    results["Section 3"] = S3
    if s3_out:
        for i, q in enumerate(["Q1","Q2","Q3","Q4","Q5","Q6","Q7","Q8","Q9","Q10"]):
            _write_score(s3_out, 31+i, 3, S3[q]["rate"])

    # ════════════════════════════════════════════════════════════════════
    # SECTION 4  –  grading col P (16), rows 14–21
    # ════════════════════════════════════════════════════════════════════
    s4_sol = sol_wb_d["Section 4"]
    s4_stu = _ws(stu_wb_d, "Section 4")
    s4_out = _ws(stu_wb,   "Section 4")

    S4 = {}
    rate_t, det_t = _check_table_exists(stu_wb_d, "MatchResults")
    S4["Q1"] = {"max":1, "rate":rate_t, "detail":det_t}

    if s4_stu:
        r = _match_rate(_rect_values(s4_stu,13,37,8,10), _rect_values(s4_sol,13,37,8,10))
        S4["Q2"] = {"max":1, "rate":r, "detail":f"{'✔' if r>=0.8 else '✘'} {int(r*100)}%"}
        S4["Q3"] = {"max":3, "rate":None, "detail":"⚠ MANUAL REVIEW – color scale CF"}
        S4["Q4"] = {"max":2, "rate":None, "detail":"⚠ MANUAL REVIEW – icon sets CF"}
        S4["Q5"] = {"max":3, "rate":None, "detail":"⚠ MANUAL REVIEW – slicer"}
        sv_v = s4_stu.cell(8,10).value
        rv_v = s4_sol.cell(8,10).value
        r6   = 1.0 if _eq_val(sv_v, rv_v) else 0.0
        S4["Q6"] = {"max":5, "rate":r6,
                    "detail":f"{'✔' if r6 else '✘'} Student={sv_v} | Expected={rv_v}"}
        S4["Q7"] = {"max":5, "rate":None, "detail":"⚠ MANUAL REVIEW – row highlight CF"}
        S4["Q8"] = {"max":5, "rate":None, "detail":"⚠ MANUAL REVIEW – column highlight CF"}
    else:
        for q, m in [("Q2",1),("Q3",3),("Q4",2),("Q5",3),("Q6",5),("Q7",5),("Q8",5)]:
            S4[q] = {"max":m, "rate":0.0, "detail":"✘ Sheet not found"}

    results["Section 4"] = S4
    if s4_out:
        for i, q in enumerate(["Q1","Q2","Q3","Q4","Q5","Q6","Q7","Q8"]):
            rate = S4[q]["rate"]
            _write_score(s4_out, 14+i, 16, rate if rate is not None else 0.0,
                         is_manual=(rate is None))

    # ── save ─────────────────────────────────────────────────────────────
    out_path = student_path.parent / f"GRADED_{student_path.name}"
    stu_wb.save(str(out_path))

    # ── summary ───────────────────────────────────────────────────────────
    summary = {}
    for sec, data in results.items():
        auto_qs   = {q: v for q, v in data.items() if v["rate"] is not None}
        manual_qs = [q for q, v in data.items() if v["rate"] is None]
        earned    = sum(v["max"] * v["rate"] for v in auto_qs.values())
        max_auto  = sum(v["max"] for v in auto_qs.values())
        max_total = sum(v["max"] for v in data.values())
        summary[sec] = {
            "score":            round(earned, 2),
            "max_auto":         max_auto,
            "max_total":        max_total,
            "manual_questions": manual_qs,
            "questions":        data,
        }

    return {"summary": summary, "output_path": out_path}"""
Football Excel Exam Auto-Grader
Compares student submissions against the solution file and fills in grading cells.

Grading cells receive a 0.0–1.0 percentage so the existing formula =score%*max_pts
computes partial credit automatically.
"""

import shutil
import subprocess
import tempfile
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

# ── paths ──────────────────────────────────────────────────────────────────
SOLUTION_PATH = Path(__file__).parent / "solution.xlsx"
SCRIPTS_DIR   = Path(__file__).parent / "scripts"

# ── colour helpers ──────────────────────────────────────────────────────────
GREEN  = PatternFill("solid", start_color="C6EFCE")
YELLOW = PatternFill("solid", start_color="FFEB9C")
RED    = PatternFill("solid", start_color="FFC7CE")
GREEN_FONT  = Font(color="276221", bold=True)
YELLOW_FONT = Font(color="9C5700", bold=True)
RED_FONT    = Font(color="9C0006", bold=True)

NUM_TOL = 0.02   # 2% relative tolerance


def _eq_val(a, b):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        fa, fb = float(a), float(b)
        if fb == 0:
            return abs(fa) < 1e-6
        return abs(fa - fb) / abs(fb) <= NUM_TOL
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def _col_values(ws, col, row_start, row_end):
    return [ws.cell(r, col).value for r in range(row_start, row_end + 1)]


def _rect_values(ws, row_start, row_end, col_start, col_end):
    return [
        [ws.cell(r, c).value for c in range(col_start, col_end + 1)]
        for r in range(row_start, row_end + 1)
    ]


def _match_rate(student_vals, sol_vals):
    """Positional match rate 0.0–1.0."""
    if not student_vals or not sol_vals:
        return 0.0
    if isinstance(student_vals[0], list):
        flat_s = [v for row in student_vals for v in row]
        flat_r = [v for row in sol_vals     for v in row]
    else:
        flat_s, flat_r = list(student_vals), list(sol_vals)
    pairs = list(zip(flat_s, flat_r))
    if not pairs:
        return 0.0
    return sum(_eq_val(a, b) for a, b in pairs) / len(pairs)


def _set_match_rate(student_vals, sol_vals):
    """
    Order-independent match for dynamic arrays (FILTER/VSTACK).
    Returns fraction of solution values found anywhere in student values.
    """
    def norm(v):
        if v is None:
            return None
        try:
            return round(float(v), 2)
        except (TypeError, ValueError):
            return str(v).strip().lower()

    flat_s = [norm(v) for row in student_vals for v in row if v is not None]
    flat_r = [norm(v) for row in sol_vals     for v in row if v is not None]
    if not flat_r:
        return 0.0
    stu_pool = list(flat_s)
    matched  = 0
    for val in flat_r:
        if val in stu_pool:
            matched += 1
            stu_pool.remove(val)
    return matched / len(flat_r)


def _write_score(ws, row, col, rate, is_manual=False):
    """Write 0.0–1.0 into a grading cell with colour coding."""
    cell = ws.cell(row, col)
    if is_manual:
        cell.value = None
        cell.fill  = YELLOW
        cell.font  = YELLOW_FONT
    else:
        cell.value = round(rate, 4)
        if rate >= 0.8:
            cell.fill = GREEN;  cell.font = GREEN_FONT
        elif rate >= 0.4:
            cell.fill = YELLOW; cell.font = YELLOW_FONT
        else:
            cell.fill = RED;    cell.font = RED_FONT


def recalculate(path: Path) -> Path:
    tmp = Path(tempfile.mkdtemp()) / path.name
    shutil.copy(path, tmp)
    try:
        result = subprocess.run(
            ["python3", str(SCRIPTS_DIR / "recalc.py"), str(tmp)],
            capture_output=True, text=True, timeout=60,
            cwd=str(SCRIPTS_DIR.parent)
        )
        if result.returncode != 0:
            print(f"[WARN] recalc: {result.stderr[:200]}")
    except Exception as e:
        print(f"[WARN] recalc failed: {e}")
    return tmp


def _check_named_ranges(wb, expected_names):
    defined = {n.lower() for n in wb.defined_names.keys()}
    missing = [n for n in expected_names if n.lower() not in defined]
    found   = len(expected_names) - len(missing)
    rate    = found / len(expected_names)
    if not missing:
        return 1.0, "✔ All named ranges found"
    return rate, f"✘ {found}/{len(expected_names)} – Missing: {', '.join(missing)}"


def _check_table_exists(wb, table_name):
    for ws in wb.worksheets:
        for tbl in ws.tables.values():
            if tbl.name.lower() == table_name.lower():
                return 1.0, f"✔ Table '{table_name}' found"
    return 0.0, f"✘ Table '{table_name}' not found"


def grade_file(student_path: Path, recalc: bool = True) -> dict:
    if recalc:
        calc_path = recalculate(student_path)
        sol_path  = recalculate(SOLUTION_PATH)
    else:
        calc_path = student_path
        sol_path  = SOLUTION_PATH

    sol_wb_d = load_workbook(str(sol_path),  data_only=True)
    stu_wb_d = load_workbook(str(calc_path), data_only=True)
    stu_wb   = load_workbook(str(calc_path))  # write scores onto recalculated copy

    results = {}

    # ════════════════════════════════════════════════════════════════════
    # SECTION 1  –  grading col Y (25), rows 3–13
    # ════════════════════════════════════════════════════════════════════
    s1_sol = sol_wb_d["Section 1 "]
    # fix .get for openpyxl workbooks
    def _ws(wb, name):
        return wb[name] if name in wb.sheetnames else None

    s1_stu = _ws(stu_wb_d, "Section 1 ")
    s1_out = _ws(stu_wb,   "Section 1 ")

    S1 = {}
    rate0, det0 = _check_named_ranges(stu_wb_d,
        ["PlayerName","Goals","Assists","Matches","Salary",
         "MarketValue","YellowCards","RedCards","Position"])
    S1["Q0"] = {"max":1, "rate":rate0, "detail":det0}

    if s1_stu:
        for label, max_pts, col in [
            ("Q1",1,12),("Q2",2,13),("Q3",2,14),("Q4",2,15),("Q5",5,16),
            ("Q6",2,17),("Q7",5,18),("Q8",5,19),("Q9",5,20),("Q10",5,21)]:
            r = _match_rate(_col_values(s1_stu,col,16,27), _col_values(s1_sol,col,16,27))
            S1[label] = {"max":max_pts, "rate":r,
                         "detail":f"{'✔' if r>=0.8 else '✘'} {int(r*100)}% cells correct"}
    else:
        for label, m in [("Q1",1),("Q2",2),("Q3",2),("Q4",2),("Q5",5),
                          ("Q6",2),("Q7",5),("Q8",5),("Q9",5),("Q10",5)]:
            S1[label] = {"max":m, "rate":0.0, "detail":"✘ Sheet not found"}

    results["Section 1"] = S1
    if s1_out:
        for i, q in enumerate(["Q0","Q1","Q2","Q3","Q4","Q5","Q6","Q7","Q8","Q9","Q10"]):
            _write_score(s1_out, 3+i, 25, S1[q]["rate"])

    # ════════════════════════════════════════════════════════════════════
    # SECTION 2  –  grading col O (15), rows 3–12
    # ════════════════════════════════════════════════════════════════════
    s2_sol = sol_wb_d["Section 2"]
    s2_stu = _ws(stu_wb_d, "Section 2")
    s2_out = _ws(stu_wb,   "Section 2")

    S2 = {}
    if s2_stu:
        # Q1 – Team/League/Stadium (cols C-E, rows 15-22)
        r = _match_rate(_rect_values(s2_stu,15,22,3,5), _rect_values(s2_sol,15,22,3,5))
        S2["Q1"] = {"max":1, "rate":r, "detail":f"{'✔' if r>=0.8 else '✘'} {int(r*100)}%"}

        # Q2 – Capacity col F
        r = _match_rate(_col_values(s2_stu,6,15,22), _col_values(s2_sol,6,15,22))
        S2["Q2"] = {"max":2, "rate":r, "detail":f"{'✔' if r>=0.8 else '✘'} {int(r*100)}%"}

        # Q3 – Ticket Price col G
        r = _match_rate(_col_values(s2_stu,7,15,22), _col_values(s2_sol,7,15,22))
        S2["Q3"] = {"max":2, "rate":r, "detail":f"{'✔' if r>=0.8 else '✘'} {int(r*100)}%"}

        # Q4 – Coach + Revenue cols H-I
        r = _match_rate(_rect_values(s2_stu,15,22,8,9), _rect_values(s2_sol,15,22,8,9))
        S2["Q4"] = {"max":1, "rate":r, "detail":f"{'✔' if r>=0.8 else '✘'} {int(r*100)}%"}

        # Q5 – Attendance % col J
        r = _match_rate(_col_values(s2_stu,10,15,22), _col_values(s2_sol,10,15,22))
        S2["Q5"] = {"max":1, "rate":r, "detail":f"{'✔' if r>=0.8 else '✘'} {int(r*100)}%"}

        # Q6 – Drop-downs
        try:
            has_dv = len(list(s2_stu.data_validations.dataValidation)) >= 2
            S2["Q6"] = {"max":1, "rate":1.0 if has_dv else 0.0,
                        "detail":"✔ Drop-downs detected" if has_dv else "✘ Drop-downs missing"}
        except Exception:
            S2["Q6"] = {"max":1, "rate":0.0, "detail":"⚠ Could not check"}

        # Q7 – manual
        S2["Q7"] = {"max":5, "rate":None, "detail":"⚠ MANUAL REVIEW – conditional formatting"}

        # Q8 – FILTER+CHOOSECOLS: formula in B32, results spill B32:C39
        # Compare set of (team_name, revenue) pairs
        sv8 = _rect_values(s2_stu, 32, 39, 2, 3)
        rv8 = _rect_values(s2_sol, 32, 39, 2, 3)
        r8  = _set_match_rate(sv8, rv8)
        S2["Q8"] = {"max":5, "rate":r8,
                    "detail":f"{'✔' if r8>=0.8 else '✘'} {int(r8*100)}% (FILTER+CHOOSECOLS, order-independent)"}

        # Q9 – LET+FILTER: formula in E32, results spill E32:F35 (4 rows)
        # Extend range to catch different-length results
        sv9 = _rect_values(s2_stu, 32, 39, 5, 6)
        rv9 = _rect_values(s2_sol, 32, 35, 5, 6)
        r9  = _set_match_rate(sv9, rv9)
        S2["Q9"] = {"max":5, "rate":r9,
                    "detail":f"{'✔' if r9>=0.8 else '✘'} {int(r9*100)}% (LET+FILTER, order-independent)"}

        # Q10 – VSTACK: formula in H32, results spill H32:I35 (4 rows)
        sv10 = _rect_values(s2_stu, 32, 39, 8, 9)
        rv10 = _rect_values(s2_sol, 32, 35, 8, 9)
        r10  = _set_match_rate(sv10, rv10)
        S2["Q10"] = {"max":5, "rate":r10,
                     "detail":f"{'✔' if r10>=0.8 else '✘'} {int(r10*100)}% (VSTACK, order-independent)"}

        # Q11-Q14 – multi-criteria, col E rows 47-50
        for label, max_pts, row in [("Q11",2,47),("Q12",5,48),("Q13",5,49),("Q14",5,50)]:
            sv_v = s2_stu.cell(row,5).value
            rv_v = s2_sol.cell(row,5).value
            r    = 1.0 if _eq_val(sv_v, rv_v) else 0.0
            S2[label] = {"max":max_pts, "rate":r,
                         "detail":f"{'✔' if r else '✘'} Student={sv_v} | Expected={rv_v}"}
    else:
        for q, m in [("Q1",1),("Q2",2),("Q3",2),("Q4",1),("Q5",1),("Q6",1),
                     ("Q7",5),("Q8",5),("Q9",5),("Q10",5),
                     ("Q11",2),("Q12",5),("Q13",5),("Q14",5)]:
            S2[q] = {"max":m, "rate":0.0, "detail":"✘ Sheet not found"}

    results["Section 2"] = S2
    if s2_out:
        for i, q in enumerate(["Q1","Q2","Q3","Q4","Q5","Q6","Q7","Q8","Q9","Q10"]):
            rate = S2[q]["rate"]
            _write_score(s2_out, 3+i, 15, rate if rate is not None else 0.0,
                         is_manual=(rate is None))
        for i, q in enumerate(["Q11","Q12","Q13","Q14"]):
            _write_score(s2_out, 47+i, 10, S2[q]["rate"])

    # ════════════════════════════════════════════════════════════════════
    # SECTION 3  –  grading col C (3), rows 31–40
    # ════════════════════════════════════════════════════════════════════
    s3_sol = sol_wb_d["Section 3"]
    s3_stu = _ws(stu_wb_d, "Section 3")
    s3_out = _ws(stu_wb,   "Section 3")

    S3 = {}
    if s3_stu:
        for label, max_pts, col in [
            ("Q1",1,7),("Q2",1,8),("Q3",2,9),("Q4",2,10),("Q5",2,11),
            ("Q8",5,14),("Q9",5,15),("Q10",5,16)]:
            r = _match_rate(_col_values(s3_stu,col,17,26), _col_values(s3_sol,col,17,26))
            S3[label] = {"max":max_pts, "rate":r,
                         "detail":f"{'✔' if r>=0.8 else '✘'} {int(r*100)}% cells correct"}

        # Q6 – Age: DATEDIF(birthdate, TODAY(), "Y")
        # Compute expected ages directly from birth dates so we never depend on
        # which date the solution file was last saved / recalculated.
        from datetime import date
        today = date.today()
        expected_ages = []
        for r in range(17, 27):
            bday = s3_stu.cell(r, 5).value   # col E = Birth Date
            if bday is not None:
                try:
                    bd = bday.date() if hasattr(bday, "date") else bday
                    age = today.year - bd.year - ((today.month, today.day) < (bd.month, bd.day))
                    expected_ages.append(age)
                except Exception:
                    expected_ages.append(None)
            else:
                expected_ages.append(None)
        student_ages = _col_values(s3_stu, 12, 17, 26)
        # Allow ±1 tolerance (birthday edge cases)
        def _age_ok(s, e):
            if s is None or e is None:
                return False
            try:
                return abs(int(float(s)) - int(e)) <= 1
            except (TypeError, ValueError):
                return False
        if expected_ages:
            matched = sum(_age_ok(s, e) for s, e in zip(student_ages, expected_ages))
            rate6   = matched / len(expected_ages)
        else:
            rate6 = 0.0
        S3["Q6"] = {"max":5, "rate":rate6,
                    "detail":f"{'✔' if rate6>=0.8 else '✘'} {int(rate6*100)}% ages correct (computed from birth dates)"}

        # Q7 – Full Name: PROPER(LastName) & ", " & PROPER(FirstName)
        # Build expected values directly from first/last name columns in the student file
        # so we don't depend on whether the student's formula recalculated.
        expected_names = []
        for r in range(17, 27):
            first = s3_stu.cell(r, 2).value   # col B = First Name
            last  = s3_stu.cell(r, 3).value   # col C = Last Name
            if first is not None and last is not None:
                expected_names.append(f"{str(last).strip().title()}, {str(first).strip().title()}")
            else:
                expected_names.append(None)
        student_names = _col_values(s3_stu, 13, 17, 26)
        def _name_ok(s, e):
            if s is None or e is None:
                return False
            return str(s).strip().lower() == str(e).strip().lower()
        if expected_names:
            matched7 = sum(_name_ok(s, e) for s, e in zip(student_names, expected_names))
            rate7    = matched7 / len(expected_names)
        else:
            rate7 = 0.0
        S3["Q7"] = {"max":2, "rate":rate7,
                    "detail":f"{'✔' if rate7>=0.8 else '✘'} {int(rate7*100)}% full names correct (LastName, FirstName format)"}
    else:
        for q, m in [("Q1",1),("Q2",1),("Q3",2),("Q4",2),("Q5",2),("Q6",5),
                     ("Q7",2),("Q8",5),("Q9",5),("Q10",5)]:
            S3[q] = {"max":m, "rate":0.0, "detail":"✘ Sheet not found"}

    results["Section 3"] = S3
    if s3_out:
        for i, q in enumerate(["Q1","Q2","Q3","Q4","Q5","Q6","Q7","Q8","Q9","Q10"]):
            _write_score(s3_out, 31+i, 3, S3[q]["rate"])

    # ════════════════════════════════════════════════════════════════════
    # SECTION 4  –  grading col P (16), rows 14–21
    # ════════════════════════════════════════════════════════════════════
    s4_sol = sol_wb_d["Section 4"]
    s4_stu = _ws(stu_wb_d, "Section 4")
    s4_out = _ws(stu_wb,   "Section 4")

    S4 = {}
    rate_t, det_t = _check_table_exists(stu_wb_d, "MatchResults")
    S4["Q1"] = {"max":1, "rate":rate_t, "detail":det_t}

    if s4_stu:
        r = _match_rate(_rect_values(s4_stu,13,37,8,10), _rect_values(s4_sol,13,37,8,10))
        S4["Q2"] = {"max":1, "rate":r, "detail":f"{'✔' if r>=0.8 else '✘'} {int(r*100)}%"}
        S4["Q3"] = {"max":3, "rate":None, "detail":"⚠ MANUAL REVIEW – color scale CF"}
        S4["Q4"] = {"max":2, "rate":None, "detail":"⚠ MANUAL REVIEW – icon sets CF"}
        S4["Q5"] = {"max":3, "rate":None, "detail":"⚠ MANUAL REVIEW – slicer"}
        sv_v = s4_stu.cell(8,10).value
        rv_v = s4_sol.cell(8,10).value
        r6   = 1.0 if _eq_val(sv_v, rv_v) else 0.0
        S4["Q6"] = {"max":5, "rate":r6,
                    "detail":f"{'✔' if r6 else '✘'} Student={sv_v} | Expected={rv_v}"}
        S4["Q7"] = {"max":5, "rate":None, "detail":"⚠ MANUAL REVIEW – row highlight CF"}
        S4["Q8"] = {"max":5, "rate":None, "detail":"⚠ MANUAL REVIEW – column highlight CF"}
    else:
        for q, m in [("Q2",1),("Q3",3),("Q4",2),("Q5",3),("Q6",5),("Q7",5),("Q8",5)]:
            S4[q] = {"max":m, "rate":0.0, "detail":"✘ Sheet not found"}

    results["Section 4"] = S4
    if s4_out:
        for i, q in enumerate(["Q1","Q2","Q3","Q4","Q5","Q6","Q7","Q8"]):
            rate = S4[q]["rate"]
            _write_score(s4_out, 14+i, 16, rate if rate is not None else 0.0,
                         is_manual=(rate is None))

    # ── save ─────────────────────────────────────────────────────────────
    out_path = student_path.parent / f"GRADED_{student_path.name}"
    stu_wb.save(str(out_path))

    # ── summary ───────────────────────────────────────────────────────────
    summary = {}
    for sec, data in results.items():
        auto_qs   = {q: v for q, v in data.items() if v["rate"] is not None}
        manual_qs = [q for q, v in data.items() if v["rate"] is None]
        earned    = sum(v["max"] * v["rate"] for v in auto_qs.values())
        max_auto  = sum(v["max"] for v in auto_qs.values())
        max_total = sum(v["max"] for v in data.values())
        summary[sec] = {
            "score":            round(earned, 2),
            "max_auto":         max_auto,
            "max_total":        max_total,
            "manual_questions": manual_qs,
            "questions":        data,
        }

    return {"summary": summary, "output_path": out_path}
